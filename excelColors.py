import openpyxl 
from openpyxl.styles import PatternFill


def main():
    book = openpyxl.load_workbook('pract.xlsx')
    sheet = book.get_sheet_by_name('Sheet2')
    column_pos = 'G1'
    warnlist = openpyxl.load_workbook('warninglist.xlsx')
    warnlistSheet = warnlist.get_sheet_by_name('Sheet1')
    # print("rows:",warnlistSheet.max_row)

    # list_val = [3,4,14,24,35,41]
    # list_dict = {3:"COVERED" , 4:"COVERED" , 14:"NA",24:"Atlas",35:"Covered",41:"NA as per specs"}
    # list_val = list_val.sort()

    # print(list_val)
    
    sheet[column_pos].value = "REQ_ID"
    
    print(sheet['G1'].value)

    print("Rows:",sheet.max_row)
    print("Columns:",sheet.max_column)

    # print("val",sheet.cell(row=3,column=1).value,sheet.cell(row=4,column=1).value,sheet.cell(row=5,column=1).value,sheet.cell(row=6,column=1).value)

    for rows in range(1,sheet.max_row+1):
        for columns in range(1,sheet.max_column):
            sheet.cell(row=rows,column=columns).fill = PatternFill(start_color="00ff00",end_color="00ff00",fill_type="solid")

    sheet[column_pos].value = "REQ_ID"

    i=2
    id = 1
    string = "REQ_GML_WRN_000"
    missed_warning = []
    for warn in range(1,warnlistSheet.max_row+1):
        NoWarn = 0
        for rows in range(1,sheet.max_row+1):
            if warnlistSheet.cell(row= warn,column=1).value == sheet.cell(row=rows,column=1).value:
                for columns in range(1,sheet.max_column):            
                    if warnlistSheet.cell(row=warn,column=2).value=="" or warnlistSheet.cell(row=warn,column=2).value.upper() == "COVERED":
                        sheet.cell(row=rows,column=columns).fill = PatternFill(start_color="40e0d0",end_color="40e0d0",fill_type="solid")
                        if columns == sheet.max_column-1:
                            # print("max1",id,warnlistSheet.cell(row= warn,column=1).value)
                            sheet.cell(row=rows,column=columns+1).value = string + str(id)
                            id=id+1
                            NoWarn = 1
                            
                    else:
                        sheet.cell(row=rows,column=columns).fill = PatternFill(start_color="a7a895",end_color="a7a895",fill_type="solid")
                        if columns == sheet.max_column-1:
                            # print("max2",id,warnlistSheet.cell(row= warn,column=1).value)
                            sheet.cell(row=rows,column=columns+1).value = warnlistSheet.cell(row=warn,column=2).value
                            NoWarn = 1


               
                            
        # if NoWarn == 0:
        #     missed_warning.append(warnlistSheet.cell(row=warn,column=1).value)

        # msd_warn.append(warnlistSheet.cell(row= warn,column=1).value if NoWarn == 0 else "" )            

    # print("missed warnings::",missed_warning)

            # if i< warnlistSheet.max_row+1 and warnlistSheet.cell(row=i,column=1).value == sheet.cell(row=rows,column=1).value:
    #             sheet.cell(row=rows,column=columns).fill = PatternFill(start_color="40e0d0",end_color="40e0d0",fill_type="solid")
    #             print("rows:",type(rows),columns)
    #             # print("max column:",sheet.max_column)
    #             if columns == sheet.max_column -1 :
    #                 # print("Columns:",columns)
    #                 # print("dict:",list_dict.get(list_val[i]))

    #                 # if list_dict.get(list_val[i]).upper() == "COVERED":
    #                 if warnlistSheet.cell(row=i,column=2).value=="" or warnlistSheet.cell(row=i,column=2).value.upper() == "COVERED":
    #                     sheet.cell(row=rows,column=columns+1).value = string + str(id)
    #                     id=id+1
    #                 else:
    #                     sheet.cell(row=rows,column=columns+1).value = warnlistSheet.cell(row=i,column=2).value
                    
    #                 # string = string + 1
    #                 i=i+1

    #         else:
    #             sheet.cell(row=rows,column=columns).fill = PatternFill(start_color="008000",end_color="008000",fill_type="solid")
            
            # print(sheet.cell(row=rows,column=columns).value)



    book.save('pract.xlsx')


if __name__ == '__main__':
    main()    

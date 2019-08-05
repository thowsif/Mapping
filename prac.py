import openpyxl
from openpyxl.styles import PatternFill

book = openpyxl.load_workbook('pract.xlsx')
sheet = book.get_sheet_by_name('Sheet2')





for warn in range(1,warnlistSheet.max_row+1):
        for rows in range(1,sheet.max_row+1):
            for columns in range(1,sheet.max_column):
                if warnlistSheet.cell(row=i,column=1).value == sheet.cell(row=rows,column=1).value:
                    if warnlistSheet.cell(row=i,column=2).value=="" or warnlistSheet.cell(row=i,column=2).value.upper() == "COVERED":
                        sheet.cell(row=rows,column=columns+1).value = string + str(id)
                        id=id+1
                    else:
                        sheet.cell(row=rows,column=columns+1).value = warnlistSheet.cell(row=i,column=2).value